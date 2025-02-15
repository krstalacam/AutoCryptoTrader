import sys
import time  # Zaman ölçümü için time modülü
import pandas as pd
import ccxt
import threading
from dash import Dash, dcc, html, Input, Output
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from cripto_bot_v1.inducatorv_main.database_manager import DatabaseManager
from cripto_bot_v1.inducatorv_main.indicators.indicator_calculator import IndicatorCalculator
from cripto_bot_v1.inducatorv_main.plot import Plotter

class App:
    def __init__(self, db_path, crypto_symbol, display_mode, result_collector):
        self.stop_thread = False
        self.db_manager = DatabaseManager(db_path)
        self.indicator_calculator = IndicatorCalculator()
        self.data = pd.DataFrame()  # DataFrame for the crypto
        self.crypto_symbol = crypto_symbol
        self.display_mode = display_mode
        self.result_collector = result_collector
        self.initial_balance = 100

    def run_console_mode(self):
        interval_count = 0
        previous_profit = None

        while not self.stop_thread:
            num_data_points = 1
            new_data = self.db_manager.fetch_data(self.crypto_symbol, interval_count, num_data_points)

            if new_data.empty:
                print(f"No more data for {self.crypto_symbol}. Stopping thread.")
                break

            self.data = pd.concat([
                self.data, new_data
            ]).drop_duplicates(subset=['timestamp']).reset_index(drop=True)

            self.indicator_calculator.execute_trade(self.data, self.initial_balance)

            current_profit = self.indicator_calculator.total_profit_loss

            if previous_profit is None or current_profit != previous_profit:
                timestamp_value = new_data['timestamp'].iloc[0]
                self.result_collector.append({
                    'Timestamp': timestamp_value,
                    'Crypto Symbol': self.crypto_symbol,
                    'Total Profit': current_profit
                })
                previous_profit = current_profit

            if interval_count % 100 == 0:
                print(f"{self.crypto_symbol}: Intervals {interval_count}, Total Profit: {current_profit:.2f}")

            interval_count += num_data_points


def main():
    start_time = time.time()
    crypto_count = input(
        "Kaç tane kripto gireceksiniz? (all yazarsanız hepsi kontrol edilir) range ise aralık belirtir: ").upper()
    display_mode = int(input("Veriyi nasıl görüntülemek istiyorsunuz? (1: Grafik, 0: Sadece Yazdır): ").strip())

    db_path = '../crypto_data.db'

    if crypto_count == 'ALL':
        crypto_symbols = DatabaseManager(db_path).get_all_crypto_symbols(enabled_active=False)
    elif crypto_count == 'RANGE':
        range_input = [int(input(f"{i + 1}. Aralık değerini girin (başlangıç ve kaç tane olacağı): ")) for i in range(2)]
        crypto_symbols = DatabaseManager(db_path).get_crypto_symbols_in_range(range_input[0], range_input[1])
    else:
        crypto_symbols = [input(f"{i + 1}. Kripto sembolü girin: ").upper() for i in range(int(crypto_count))]

    result_collector = []
    threads = []

    for crypto_symbol in crypto_symbols:
        app_instance = App(db_path, crypto_symbol, display_mode, result_collector)
        thread = threading.Thread(target=app_instance.run_console_mode)
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    # Save results to Excel
    excel_file = 'crypto_trading_results.xlsx'
    results_df = pd.DataFrame(result_collector)
    pivot_df = results_df.pivot(index='Timestamp', columns='Crypto Symbol', values='Total Profit')
    pivot_df.reset_index(inplace=True)
    pivot_df['Total'] = pivot_df.iloc[:, 1:].sum(axis=1)
    pivot_df.to_excel(excel_file, index=False)

    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    worksheet.column_dimensions['A'].width = 18.5

    # Save workbook
    workbook.save(excel_file)
    print("Results saved to 'crypto_trading_results.xlsx'.")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Program toplamda {elapsed_time:.2f} saniye sürdü.")

if __name__ == "__main__":
    main()
