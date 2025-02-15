import sys
import time  # Zaman ölçümü için time modülü

import pandas as pd
import ccxt
import threading
from dash import Dash, dcc, html, Input, Output
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from cripto_bot_v1.inducatorv_main.database_manager import DatabaseManager
from cripto_bot_v1.inducatorv_main.indicators.indicator_calculator import IndicatorCalculator
from cripto_bot_v1.inducatorv_main.plot import Plotter


class App:
    def __init__(self, db_path, crypto_symbol, display_mode, result_collector, port=None):
        self.stop_thread = False
        self.db_manager = DatabaseManager(db_path)
        self.indicator_calculator = IndicatorCalculator()
        self.plotter = Plotter(crypto_symbol)  # crypto_symbol parametresi burada veriliyor
        self.initial_balance = 100
        self.data = pd.DataFrame()  # Global DataFrame for price data
        self.crypto_symbol = crypto_symbol
        self.display_mode = display_mode
        self.port = port
        self.exchange = ccxt.binance()
        self.result_collector = result_collector  # To collect results for all cryptos

        # Dash app initialization for graphical mode
        if display_mode == 1:
            self.app = Dash(__name__)
            self.app.layout = html.Div([
                html.Div(id='profit-loss-output',
                         style={'font-size': '30px', 'font-weight': 'bold', 'text-align': 'center',
                                'margin-bottom': '10px'}),
                dcc.Graph(id='graph', style={'width': '100%', 'height': '80vh'}),
                dcc.Interval(id='interval-component', interval=90, n_intervals=0)  # Update every minute
            ])

            @self.app.callback(
                Output('graph', 'figure'),
                Input('interval-component', 'n_intervals')
            )
            def update_graph(n_intervals):
                return self.update_graph_data()

    def update_graph_data(self):
        new_data = self.db_manager.fetch_data(self.crypto_symbol, len(self.data), 1)

        if new_data.empty:
            if not self.stop_thread:
                print(f"No more data for {self.crypto_symbol}")
            self.stop_thread = True
            return self.plotter.plot_data(
                self.data,
                self.indicator_calculator.trade_signals,
                self.indicator_calculator.total_profit_loss,
                self.indicator_calculator.trade_profit_loss
            )

        self.data = pd.concat([self.data, new_data]).drop_duplicates(subset=['timestamp']).reset_index(drop=True)

        self.indicator_calculator.execute_trade(self.data, self.initial_balance)

        # Grafik çizimi
        fig = self.plotter.plot_data(
            self.data,
            self.indicator_calculator.trade_signals,
            self.indicator_calculator.total_profit_loss,
            self.indicator_calculator.trade_profit_loss
        )

        return fig

    def run_console_mode(self):
        self.stop_thread = False  # Thread durumu için bir kontrol bayrağı
        interval_count = 0
        previous_profit = None  # To track the last profit value

        # Fetch initial data to determine total intervals for the cryptocurrency
        # initial_data = self.db_manager.fetch_data(self.crypto_symbol, 0,1_000_000)  # Adjust the number to fetch all data
        initial_data = self.db_manager.fetch_data(self.crypto_symbol, len(self.data), 1)
        # initial_data üstteki ikiside aynı çalışıyor
        total_intervals = len(initial_data)  # Total number of available intervals

        while not self.stop_thread:  # Thread durma koşulunu kontrol et
            num_data_points = 1  # Update with 1 minute data
            new_data = self.db_manager.fetch_data(self.crypto_symbol, interval_count, num_data_points)

            if new_data.empty:
                print(f"No more data for {self.crypto_symbol}. Stopping thread.")
                self.stop_thread = True  # Veri yoksa thread'i durdur
                break  # Döngüyü bitir

            self.data = pd.concat([self.data, new_data]).drop_duplicates(subset=['timestamp']).reset_index(drop=True)
            # print(f"First timestamp: {self.data.iloc[0]['timestamp']} | Last timestamp: {self.data.iloc[-1]['timestamp']}")
            self.indicator_calculator.execute_trade(self.data, self.initial_balance)

            current_profit = self.indicator_calculator.total_profit_loss

            # Only collect results if the profit has changed
            if previous_profit is None or current_profit != previous_profit:
                # Extract the timestamp from the new_data directly
                timestamp_value = new_data['timestamp'].iloc[0]  # Get the timestamp from new_data
                self.result_collector.append({
                    'Timestamp': timestamp_value,  # Use the correct timestamp value here
                    'Crypto Symbol': self.crypto_symbol,
                    'Total Profit': current_profit
                })
                previous_profit = current_profit  # Update the previous profit to the current one

            # Print status every 100 intervals
            if interval_count % 100 == 0:
                print(
                    f"Intervals {interval_count}/{total_intervals}, Total Profit for {self.crypto_symbol}: {current_profit:.2f}")

            interval_count += num_data_points  # Increment interval_count after each update

    def run_server(self):
        if self.display_mode == 1:
            # Start the Dash app without the reloader
            self.app.run_server(port=self.port, debug=True, use_reloader=False)  # Disable reloader


def main():
    start_time = time.time()  # Program başlangıcında zaman damgası al
    crypto_count = input(
        "Kaç tane kripto gireceksiniz? (all yazarsanız hepsi kontrol edilir) range ise aralik belirtir: ").upper()
    display_mode = int(input("Veriyi nasıl görüntülemek istiyorsunuz? (1: Grafik, 0: Sadece Yazdır): ").strip())

    db_path = '../crypto_data.db'

    if crypto_count == 'ALL':
        crypto_symbols = DatabaseManager(db_path).get_all_crypto_symbols(
            enabled_active=False)  # Tüm kripto sembollerini al
    elif crypto_count == 'RANGE':
        range_input = [int(input(f"{i + 1}. Aralık değerini girin (başlangıç ve kaç tane olacağı): ")) for i in range(2)]
        crypto_symbols = DatabaseManager(db_path).get_crypto_symbols_in_range(range_input[0], range_input[
            1])  # Tüm kripto sembollerini al
    else:
        crypto_symbols = [input(f"{i + 1}. Kripto sembolü girin: ").upper() for i in range(int(crypto_count))]

    app_instances = []
    threads = []
    result_collector = []  # List to collect results for all cryptocurrencies

    for idx, crypto_symbol in enumerate(crypto_symbols):
        port = 8052 + idx if display_mode == 1 else None  # Her app'e grafik modunda farklı port ver
        app_instance = App(db_path, crypto_symbol, display_mode, result_collector, port)
        app_instances.append(app_instance)

        if display_mode == 1:
            thread = threading.Thread(target=app_instance.run_server)  # Dash sunucusunu thread içinde çalıştır
        else:
            thread = threading.Thread(target=app_instance.run_console_mode)  # Konsol modunu thread içinde çalıştır

        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    if display_mode == 0:
        # Excel save operation
        excel_file = 'crypto_trading_results.xlsx'

        # Assuming 'result_collector' and 'pivot_df' already exist
        results_df = pd.DataFrame(result_collector)
        pivot_df = results_df.pivot(index='Timestamp', columns='Crypto Symbol', values='Total Profit')
        pivot_df.reset_index(inplace=True)
        pivot_df['Total'] = pivot_df.iloc[:, 1:].sum(axis=1)
        pivot_df.to_excel(excel_file, index=False)

        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        worksheet.column_dimensions['A'].width = 18.5

        # Find min and max values for conditional formatting
        min_value = float('inf')
        max_value = float('-inf')

        for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        max_value = max(max_value, cell.value)
                    elif cell.value < 0:
                        min_value = min(min_value, cell.value)

        if min_value == float('inf'):
            min_value = -100
        if max_value == float('-inf'):
            max_value = 100

        # Apply conditional formatting
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        intensity = 255 - int(255 * (cell.value / max_value)) if max_value != 0 else 0
                        cell.fill = PatternFill(start_color=f'00{intensity:02X}FF00',
                                                end_color=f'00{intensity:02X}FF00', fill_type='solid')
                    elif cell.value < 0:
                        intensity = 255 - int(255 * (abs(cell.value) / abs(min_value))) if min_value != 0 else 0
                        cell.fill = PatternFill(start_color=f'00FF{intensity:02X}00',
                                                end_color=f'00FF{intensity:02X}00', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

        # Add advanced SUM formula to the 'Total' column
        for row_num in range(2, worksheet.max_row + 1):
            last_col = worksheet.max_column
            col_letters = [chr(64 + col_num) for col_num in range(2, last_col)]
            formula_parts = []
            for col in col_letters:
                find_value_formula = f"IF(ISNUMBER({col}{row_num}), {col}{row_num}, "
                find_value_formula += f"IF(ISNUMBER({col}{row_num - 1}), {col}{row_num - 1}, "
                find_value_formula += f"IF(ISNUMBER({col}{row_num - 2}), {col}{row_num - 2}, 0)))"
                formula_parts.append(find_value_formula)
            worksheet.cell(row=row_num, column=last_col).value = f"=SUM({','.join(formula_parts)})"

        # Freeze the first row
        worksheet.freeze_panes = 'A2'

        # Save the workbook
        workbook.save(excel_file)
        print(
            "All results saved to 'crypto_trading_results.xlsx' with conditional formatting and advanced SUM formula.")

    end_time = time.time()  # Program bitişinde zaman damgası al
    elapsed_time = end_time - start_time  # Geçen süreyi hesapla
    print(f"Program toplamda {elapsed_time:.2f} saniye sürdü.")  # Konsola yazdır


if __name__ == "__main__":
    main()
